import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import streamlit.components.v1 as components
import time
import zipfile

##############################################
st.set_page_config(page_title="DataComp",layout='wide',page_icon="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcSPFg0oJPlzocM4yk6K6Q0WFzR9ISo0in5E5Q&s")

# components.html(
#     """
#     <script>
#     // Locate elements
#     var decoration = window.parent.document.querySelectorAll('[data-testid="stDecoration"]')[0];
#     var sidebar = window.parent.document.querySelectorAll('[data-testid="stSidebar"]')[0];

#     // Observe sidebar size
#     function outputsize() {
#         decoration.style.left = `${sidebar.offsetWidth}px`;
#     }

#     new ResizeObserver(outputsize).observe(sidebar);

#     // Adjust sizes
#     outputsize();
#     decoration.style.height = "3.0rem";
#     decoration.style.right = "45px";

#     // Adjust image decorations
#     decoration.style.backgroundSize = "contain";
#     </script>        
#     """, width=1, height=1)



col1, _,col2,col3 = st.columns([1,1.5,8,1])

with col1:
    #st.image("logo_st.png",width=150)
    pass

hide_img_fs = '''
<style>
button[title="View fullscreen"]{
    visibility: hidden;}
</style>
'''
st.markdown(hide_img_fs, unsafe_allow_html=True)

with col2:
    #st.write('\n\n\n')
    st.write('\n')
    st.title('Data Verification Tool',anchor=False)
#     pass
# Main content
st.header("Compare Excel Reports",anchor=False)
#st.write("*restrict to excel uploads only ")
#st.write("This is the main content area.")


# with col3:

# Sidebar content


st.markdown(
        f"""
        <style>
            [data-testid="stSidebar"] {{
                background-image: url({"https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcR6lusNEaS-BAjRAg0lhjOUKbJy0DippLQzdQ&s"});
                background-repeat: no-repeat;
                padding-top: 50px;
                background-position: 0px 0px;
                background-size : 250px 120px;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )

st.markdown(
    """
    <style>
        section[data-testid="stSidebar"] {
            width: 50px !important; # Set the width to your desired value
        }
    </style>
    """,
    unsafe_allow_html=True,
)
#st.sidebar.image("images.jpg")
st.sidebar.header("Document Type")
# st.sidebar.subheader("Subheading")
#st.sidebar.text("Select the format of report")

################################################
# from streamlit_option_menu import option_menu
# from streamlit_extras.switch_page_button import switch_page

# Comapare the files
temp_path = ''
##########################

ERROR_CODES = ['#NULL!', '#DIV/0!', '#VALUE!', '#REF!', '#NAME?', '#NUM!', '#N/A']
##########################
#Colour codes
#pink = for Errors
#red = for difference less than threshold
#green = for difference greater than threshold
#Yellow = for exception where threshold calculation is not done

fill_style_red = PatternFill(start_color = "00FF0000",end_color="00FF0000",fill_type="solid")
fill_style_green = PatternFill(start_color = "00008000",end_color="00008000",fill_type="solid")
fill_style_purp = PatternFill(start_color = "00800080",end_color="00800080",fill_type="solid")
fill_style_pin = PatternFill(start_color = "003366FF",end_color="003366FF",fill_type="solid")
fill_style_yellow = PatternFill(start_color = "00FFCC00",end_color="00FFCC00",fill_type="solid")
##################################

@st.cache_data
def load_file(file):
    df_wb = openpyxl.load_workbook(file,data_only=True)
    return df_wb
###########################

def compare_sheets(wb,file1,source_dict,reference_dict,threshold_dict):
    
    output = io.BytesIO()
    src_name = ''
    if file1.name.split('.')[-1] == 'xlsx':
        src = load_file(file1)        
    else:
        df_src = load_data(file1)
        src_name= temp_path+file1.name.split('.')[0]
        # Create a func to read the df and write all sheets to xlsx
        with pd.ExcelWriter(src_name+'.xlsx') as writer:
            for k in source_dict.keys():
                df_src[k].to_excel(writer, sheet_name=k)
        src = load_file(src_name+'.xlsx')
    
    progress_text = "Comparision in progress. Please wait."
    my_bar = st.progress(0, text=progress_text)


    percent_complete = 0
    j = round(100/len(list(source_dict.keys())))

    for key in source_dict.keys():
        print("In Sheet-----",key)
#         src = openpyxl.load_workbook(temp_path+str(i) + "_src.xlsx")
#         ref = openpyxl.load_workbook(temp_path+str(i) + "_ref.xlsx")
        percent_complete = percent_complete + j
        time.sleep(0.01)
        #if percent_complete<= 100:
        my_bar.progress(percent_complete, text=progress_text)
        wb.create_sheet(str(key))
        var_threshhold = threshold_dict[key]
        for row1,row2 in zip(src[key][source_dict[key]],src[key][reference_dict[key]]):
            for cell,cell2 in zip(row1,row2):
                current_cell_value = cell.value
                cell_location = cell.coordinate
                wb[str(key)][cell_location].value = current_cell_value
                if current_cell_value != cell2.value: 
                    var_ = 0 
                    if current_cell_value in ERROR_CODES:
                        cell.fill = fill_style_pin
                        wb[str(key)][cell_location].fill = fill_style_pin
                        
                    if (isinstance(current_cell_value, int) or isinstance(current_cell_value, float)):
                        
                        try:
                            if (cell.number_format) == '0%':
                                var_ = (current_cell_value - cell2.value)*100
                                cell.value = str(round(var_,2))+"%"
                                wb[str(key)][cell_location].value = ""+ str(round(var_,2))+"%"
                            else:
                                var_ = ((current_cell_value - cell2.value)*100)/cell2.value
                                cell.value = str(round(var_,2))+"%"
                                wb[str(key)][cell_location].value = ""+ str(round(var_,2))+"%"
                                
                        except Exception as e:         
                            cell.fill = fill_style_red
                            wb[str(key)][cell_location].fill = fill_style_yellow
                            
                        if var_ < -(var_threshhold) :     
                            cell.fill = fill_style_red
                            wb[str(key)][cell_location].fill = fill_style_red
                        
                        if var_ > var_threshhold :                    
                            cell.fill = fill_style_green
                            wb[str(key)][cell_location].fill = fill_style_green

                    if isinstance(current_cell_value, str):                 
                        cell.fill = fill_style_purp
                        wb[str(key)][cell_location].fill = fill_style_purp

                else:
                    pass
    wb.save(output)
    my_bar.empty()
    return output.getvalue()
##########################
def compare_files(wb,file1,file2,df_src,df_ref,temp_path, var_threshhold,my_bar,exclude=['']):
    #file_name = "-".join(uploaded_file1.split('\\')[-1].split(" ")[:-2]) + "-Quality-Report"
    var_threshhold = int(var_threshhold)
    output = io.BytesIO()
    percent_complete = 0
    j = round(100/len(list(df_src.keys())))
    src_name = ''
    ref_name = ''
    if file1.name.split('.')[-1] == file2.name.split('.')[-1] == 'xlsx':
        src = load_file(file1)        
        ref = load_file(file2)
    else:
        src_name= temp_path+file1.name.split('.')[0]
        ref_name= temp_path+file2.name.split('.')[0]
        # Create a func to read the df and write all sheets to xlsx
        with pd.ExcelWriter(src_name+'.xlsx') as writer, pd.ExcelWriter(ref_name+'.xlsx') as writer1:
            for k in df_src.keys():
                df_src[k].to_excel(writer, sheet_name=k)
                df_ref[k].to_excel(writer1, sheet_name=k)

        src = load_file(src_name+'.xlsx')
        ref = load_file(ref_name+'.xlsx')

    for i in list(df_src.keys()):
        #df_src[i].to_excel(temp_path+str(i) + "_src.xlsx",index=False,header=False)
        #df_ref[i].to_excel(temp_path+str(i) + "_ref.xlsx",index=False,header=False)
        if i in exclude:
            continue
        print("In Sheet-----",i)
#         src = openpyxl.load_workbook(temp_path+str(i) + "_src.xlsx")
#         ref = openpyxl.load_workbook(temp_path+str(i) + "_ref.xlsx")
        percent_complete = percent_complete + j
        time.sleep(0.01)
        #if percent_complete<= 100:
        my_bar.progress(percent_complete, text=progress_text)
        wb.create_sheet(str(i))
        for row in src[i].iter_rows():
            for cell in row:
                current_cell_value = cell.value
                cell_location = cell.coordinate
                wb[str(i)][cell_location].value = current_cell_value
                #print(current_cell_value)
#                 print(cell.number_format)
#                 if (cell.number_format) == '0%':
#                     print('YES')
                
                if current_cell_value != ref[i][cell_location].value:
                    var_ = 0
                    if current_cell_value in ERROR_CODES:   
                        cell.fill = fill_style_pin
                        wb[str(i)][cell_location].fill = fill_style_pin
                        
                    if (isinstance(current_cell_value, int) or isinstance(current_cell_value, float)):
                        
                        try:
                            if (cell.number_format) == '0%':
                                var_ = (current_cell_value - ref[i][cell_location].value)*100
                                cell.value = str(round(var_,2))+"%"
                                wb[str(i)][cell_location].value = ""+ str(round(var_,2))+"%"
                            else:
                                var_ = ((current_cell_value - ref[i][cell_location].value)*100)/ref[i][cell_location].value
                                cell.value = str(round(var_,2))+"%"
                                wb[str(i)][cell_location].value = ""+ str(round(var_,2))+"%"
                                
                        except Exception as e:         
                            cell.fill = fill_style_red
                            wb[str(i)][cell_location].fill = fill_style_yellow
                            
                        if var_ < -(var_threshhold) :     
                            cell.fill = fill_style_red
                            wb[str(i)][cell_location].fill = fill_style_red
                        
                        if var_ > var_threshhold :                    
                            cell.fill = fill_style_green
                            wb[str(i)][cell_location].fill = fill_style_green

                    if isinstance(current_cell_value, str):                 
                        cell.fill = fill_style_purp
                        wb[str(i)][cell_location].fill = fill_style_purp

                else:
                    pass
                
        #wb.save(temp_path+file_name+".xlsx")
        src.close()
        ref.close()
        wb.save(output)
        try:
            os.remove(src_name+'.xlsx')
            os.remove(ref_name+'.xlsx')
        except :
            pass
    print("++++++++++++++Done+++++++++++++++++++")
    my_bar.empty()

    return output.getvalue()
################


@st.cache_data
def load_data(file):
    df = pd.read_excel(file,engine=None,sheet_name = None)
    return df


################

selection = st.sidebar.radio('Select Format', options=['Excel Report', 'Multi Excel Reports','PDF Report'])
#st.sidebar.checkbox('special')
if selection == 'Excel Report':
    col1,col2 = st.columns([5,5])
    #uploaded_file1 = None
    dataframe1 = None
    dataframe2 = None

    with col1:
        uploaded_file1 = st.file_uploader("Upload Report file for quality check")
        st.write("-------------------------------")
        #if uploaded_file1 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe1 = pd.read_excel(uploaded_file1,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe1.keys())
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")

    with col2:
        uploaded_file2 = st.file_uploader("Upload Reference Report file for Comparision")
        st.write("-------------------------------")
        #if uploaded_file2 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe2 = pd.read_excel(uploaded_file2,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe2.keys())
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")
    status = False
    excel_file = ''
    if (uploaded_file1 is not None) & (uploaded_file2 is not None):     
        dataframe1 = load_data(uploaded_file1) 
                # except:
                #     print("check fileformat")
                
        st.write("Total Sheets ",dataframe1.keys())

        dataframe2 = load_data(uploaded_file2)
        st.write("Total Sheets ",dataframe2.keys())

        ## Check for unequal sheets:
        if dataframe1.keys() == dataframe2.keys():
        
            check_list = (list(dataframe1.keys()))
            agree = st.checkbox("Do you want to select specific sheets")
            options = check_list
            if agree:
                options = st.multiselect(
                        "Select sheets",
                        check_list)
            with st.form("my_form"):
                # var_threshhold = st.select_slider("Select a threshold for the QC comparison",
                #             options=[ i for i in range(1,101,1)])

                var_threshhold = st.number_input("Enter the threshold to compare",
                                            max_value=100,
                                            step = 1,
                                            placeholder="Type a number...")
                
                submitted = st.form_submit_button("Submit",type="primary")
                if submitted:
                    # try:
                    #st.write("You selected:", options)
                    exclude = [remove_ for remove_ in check_list if remove_ not in options]
                    #st.write("You opted out :", exclude)
                    print("Threshold selcted : ",var_threshhold)
                    st.write("Threshold selcted : ",var_threshhold)
                    #st.write("Comparision in progress")

                    progress_text = "Comparision in progress. Please wait."
                    my_bar = st.progress(0, text=progress_text)

                    wb=openpyxl.Workbook()
                    wb.remove(wb['Sheet'])

                    #call the compare function to prepare the report
                    excel_file = compare_files(wb,uploaded_file1,uploaded_file2,dataframe1,dataframe2,temp_path, var_threshhold,my_bar,exclude)     
                    my_bar.empty()
                
            status = st.download_button(
                label='Export to Excel'
                , data = excel_file # excel file is the name of the downloaded <class ‘xlsxwriter.workbook.Workbook’>
                , file_name= "Quality-Report"+".xlsx"
                , disabled = not excel_file
                )
        else:
            st.write("The uploaded files are not having same sheets, please re-upload")
            
    #Add custom comparision of same sheet
        custom_compare = st.checkbox("Do you want to custom compares sheets and rows in a particular file")
        excel_file_ = ''
        if custom_compare:
            uploaded_file3 = st.file_uploader("Upload Report file for Custome Comparision")
            if uploaded_file3 is not None :     
                dataframe3 = load_data(uploaded_file3)
                check_list = (list(dataframe3.keys()))
                options = check_list
                options = st.multiselect('Please select specific sheets for comparision',
                            options=check_list)
                source_dict = {}
                reference_dict = {}
                threshold_dict = {}
                for sheet_name in options:
                    col1,col2,col3 = st.columns([3,3,3])
                    with col1:
                        tsk = st.text_input('Enter the start and end cell for source ex\: **A1\:J15** ', placeholder = sheet_name)
                        source_dict[sheet_name] = tsk
                    with col2:
                        tsk1 = st.text_input('Enter the start and end cell for reference ex\: **A1\:J15** ', placeholder = sheet_name)
                        reference_dict[sheet_name] = tsk1
                    with col3:
                        tsk2 = st.number_input('Enter threshold for '+sheet_name,1)
                        threshold_dict[sheet_name] = tsk2

                if(st.button('Submit',type="primary")):
                        wb=openpyxl.Workbook()
                        wb.remove(wb['Sheet'])
                        excel_file_ = compare_sheets(wb,uploaded_file3,source_dict,reference_dict,threshold_dict)
        
            status = st.download_button(
                label='Export to Excel'
                , data = excel_file # excel file is the name of the downloaded <class ‘xlsxwriter.workbook.Workbook’>
                , file_name= "Quality-Report_"+".xlsx"
                , disabled = not excel_file_
                )                
        ## Add a count report at the end to see the results



## New module for multi excel reports                        
elif selection == 'Multi Excel Reports':
    col1,col2 = st.columns([5,5])
    #uploaded_file1 = None
    dataframe1 = None
    dataframe2 = None

    with col1:
        uploaded_file1 = st.file_uploader("Upload Report file for quality check",accept_multiple_files=True)
        if uploaded_file1 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe1 = pd.read_excel(uploaded_file1,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe1.keys())
            st.write("-------------------------------")
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")

    with col2:
        uploaded_file2 = st.file_uploader("Upload Reference Report file for Comparision",accept_multiple_files=True)
        if uploaded_file2 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe2 = pd.read_excel(uploaded_file2,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe2.keys())
            st.write("-------------------------------")
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")
    excel_file = ''
    var_threshhold_list = False
    if (uploaded_file1 is not None) & (uploaded_file2 is not None):
        threshhold_select = st.radio(
            "Do you want to compare all the files with same threshold or different : ",
            ["Yes", "No"])
        var_threshhold_list = {}
        all_files = {}
        if threshhold_select == "Yes":
                tsk = st.number_input("Enter the threshold to compare",
                                         max_value=100,
                                         step = 1,
                                         placeholder="Type a number...")
                for file_name in uploaded_file1:
                    var_threshhold_list[file_name.name] = tsk
            
        if threshhold_select == "No":
            for file_name in uploaded_file1:
                tsk = st.text_input('Enter the threshold for :', placeholder = file_name.name)
                var_threshhold_list[file_name.name] = tsk
        
        with st.form("my_form"):
            # var_threshhold = st.select_slider("Select a threshold for the QC comparison",
            #             options=[ i for i in range(1,101,1)])
            st.write("Selected threshold :" ,var_threshhold_list)
            submitted = st.form_submit_button("Submit",type="primary",disabled=not var_threshhold_list)
            if submitted:
                # try:
                st.write("Reading Files .....")
                st.write("This might take few minutes depending on number of files !!")
                i = 1
                for file1,file2,var_threshhold in zip(uploaded_file1,uploaded_file2,var_threshhold_list.values()):
                    dataframe1 = load_data(file1) 
                    # except:
                    #     print("check fileformat")
                    dataframe2 = load_data(file2)
                    if dataframe1.keys() == dataframe2.keys():
                        st.write("Sheets in ",file1.name,"are",dataframe1.keys())
                        st.write("Sheets in ",file2.name,"are",dataframe2.keys())

                        # print("Threshold selcted : ",var_threshhold_list)
                        # st.write("Threshold selcted : ",var_threshhold_list)
                        #st.write("Comparision in progress")
                        progress_text = "Comparision in progress. Please wait."
                        my_bar = st.progress(0, text=progress_text)
                
                        wb=openpyxl.Workbook()
                        wb.remove(wb['Sheet'])

                        #call the compare function to prepare the report
                        #print("------------", uploaded_file1)
                        excel_file = compare_files(wb,file1,file2,dataframe1,dataframe2,temp_path, var_threshhold,my_bar)
                        all_files[file1.name] = excel_file
                        my_bar.empty()
                        st.write("Compared :",file1.name, "completed : ",i,"/",len(uploaded_file1))
                        i = i+1
                        st.write("----------------------------------------------------------------")
                    else:
                        st.write("Sheets in ",file1.name,"are",dataframe1.keys())
                        st.write("Sheets in ",file2.name,"are",dataframe2.keys())
                        st.write("Either files order or sheets orders are not correct, Please recheck and upload again")
                        break
        
        #if excel_file != None:
        # For multifile download:
        buf = io.BytesIO()

        with zipfile.ZipFile(buf, "x") as csv_zip:
            for file_d in all_files.keys():
                csv_zip.writestr(file_d+".xlsx", all_files[file_d])

        st.download_button(
            label="Export file",
            data=buf.getvalue(),
            file_name="Quality-Report"+".zip",
            mime="application/zip",
            disabled = not excel_file
        )


        # st.download_button(
        #     label='Export to Excel'
        #     , data = excel_file # excel file is the name of the downloaded <class ‘xlsxwriter.workbook.Workbook’>
        #     , file_name= "Quality-Report"+".xlsx"
        #     , disabled = not excel_file
        #     )

elif selection == 'PDF Report':
    col1,col2 = st.columns([5,5])
    #uploaded_file1 = None
    dataframe1 = None
    dataframe2 = None

    with col1:
        uploaded_file1 = st.file_uploader("Upload Report file for quality check")
        if uploaded_file1 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe1 = pd.read_excel(uploaded_file1,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe1.keys())
            st.write("-------------------------------")
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")

    with col2:
        uploaded_file2 = st.file_uploader("Upload Reference Report file for Comparision")
        if uploaded_file2 is not None:
            # Can be used wherever a "file-like" object is accepted:
            # dataframe2 = pd.read_excel(uploaded_file2,engine="pyxlsb",sheet_name = None)
            # st.write("Total Sheets ",dataframe2.keys())
            st.write("-------------------------------")
            # st.write("Total rows and columns in sheet wise are are x,y")
            # st.write("-------------------------------")
            # st.write("Total rows in sheet wise are ")
    st.write("Feature Development In Progress ... ")

else:
    pass


